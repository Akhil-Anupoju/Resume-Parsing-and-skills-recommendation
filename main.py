import openpyxl
from flask import Flask, request, render_template, jsonify, send_file
import os
import docx2txt
import PyPDF2
import pandas as pd
from sklearn.feature_extraction.text import TfidfVectorizer
from sklearn.metrics.pairwise import cosine_similarity
import spacy
import re
from datetime import datetime
from openpyxl.styles import Font, Alignment, PatternFill
from openpyxl.utils import get_column_letter

app = Flask(__name__)
app.config['UPLOAD_FOLDER'] = 'uploads/'
app.config['EXCEL_FOLDER'] = 'excel_exports/'

# Load spaCy model for skills extraction
nlp = spacy.load("en_core_web_sm")

# Comprehensive list of technical and soft skills
SKILL_PATTERNS = {
    # Programming Languages
    'languages': [
        'python', 'java', 'javascript', 'typescript', 'c++', 'c#', 'ruby', 'php',
        'swift', 'kotlin', 'go', 'rust', 'scala', 'perl', 'r', 'matlab'
    ],
    # Web Technologies
    'web': [
        'html', 'css', 'react', 'angular', 'vue.js', 'node.js', 'django',
        'flask', 'spring', 'asp.net', 'jquery', 'bootstrap', 'sass', 'less',
        'webpack', 'redux', 'graphql', 'rest api'
    ],
    # Databases
    'databases': [
        'sql', 'mysql', 'postgresql', 'mongodb', 'oracle', 'redis', 'elasticsearch',
        'cassandra', 'sqlite', 'dynamodb', 'mariadb'
    ],
    # Cloud & DevOps
    'cloud_devops': [
        'aws', 'azure', 'google cloud', 'docker', 'kubernetes', 'jenkins',
        'gitlab ci', 'terraform', 'ansible', 'circleci', 'maven', 'gradle'
    ],
    # Data Science & AI
    'data_ai': [
        'machine learning', 'deep learning', 'neural networks', 'data analysis',
        'pandas', 'numpy', 'scipy', 'scikit-learn', 'tensorflow', 'pytorch',
        'tableau', 'power bi', 'statistics'
    ],
    # Soft Skills
    'soft_skills': [
        'project management', 'team leadership', 'agile', 'scrum', 'communication',
        'problem solving', 'critical thinking', 'time management', 'teamwork',
        'collaboration'
    ]
}

# Combine all skills into a single list
ALL_SKILLS = [skill for category in SKILL_PATTERNS.values() for skill in category]


def extract_skills(text):
    # Convert text to lowercase for better matching
    text = text.lower()

    # Initialize set for unique skills
    skills = set()

    # Find skills using regex with word boundaries
    for skill in ALL_SKILLS:
        # Use word boundary for exact matches and handle multi-word skills
        if re.search(r'\b' + re.escape(skill) + r'\b', text):
            skills.add(skill)

    return sorted(list(skills))


def format_excel(file_path):
    # Read the Excel file
    wb = pd.read_excel(file_path, engine='openpyxl').to_excel(file_path, index=False, engine='openpyxl')
    wb = openpyxl.load_workbook(file_path)
    ws = wb.active

    # Define styles
    header_font = Font(bold=True, size=12)
    header_fill = PatternFill(start_color="CCE5FF", end_color="CCE5FF", fill_type="solid")

    # Format headers
    for cell in ws[1]:
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)

    # Auto-adjust column widths
    for column in ws.columns:
        max_length = 0
        column_letter = get_column_letter(column[0].column)

        for cell in column:
            try:
                if len(str(cell.value)) > max_length:
                    max_length = len(str(cell.value))
            except:
                pass

        # Set width with some padding
        adjusted_width = min(max_length + 2, 50)  # Max width of 50
        ws.column_dimensions[column_letter].width = adjusted_width

        # Set alignment for data cells
        for cell in column[1:]:  # Skip header
            cell.alignment = Alignment(horizontal='left', vertical='center', wrap_text=True)

    # Save the formatted workbook
    wb.save(file_path)


@app.route('/export_excel', methods=['POST'])
def export_excel():
    data = request.get_json()

    # Prepare data for Excel
    excel_data = []
    for resume in data['resumes']:
        excel_data.append({
            'Rank': resume['rank'],
            'File Name': resume['filename'],
            'Match Score (%)': resume['score'],
            'Skills': ', '.join(resume['skills'])  # Join skills with commas
        })

    df = pd.DataFrame(excel_data)

    if not os.path.exists(app.config['EXCEL_FOLDER']):
        os.makedirs(app.config['EXCEL_FOLDER'])

    timestamp = datetime.now().strftime('%Y%m%d_%H%M%S')
    excel_path = os.path.join(app.config['EXCEL_FOLDER'], f'top_resumes_{timestamp}.xlsx')

    # Export to Excel
    df.to_excel(excel_path, index=False)

    # Apply formatting
    format_excel(excel_path)

    return jsonify({'file_path': excel_path})

# ----------------------------------------------------------------

def extract_text_from_pdf(file_path):
    text = ""
    with open(file_path, 'rb') as file:
        reader = PyPDF2.PdfReader(file)
        for page in reader.pages:
            text += page.extract_text()
    return text


def extract_text_from_docx(file_path):
    return docx2txt.process(file_path)


def extract_text_from_txt(file_path):
    with open(file_path, 'r', encoding='utf-8') as file:
        return file.read()


def extract_text(file_path):
    if file_path.endswith('.pdf'):
        return extract_text_from_pdf(file_path)
    elif file_path.endswith('.docx'):
        return extract_text_from_docx(file_path)
    elif file_path.endswith('.txt'):
        return extract_text_from_txt(file_path)
    else:
        return ""


@app.route("/")
def matchresume():
    return render_template('matchresume.html')


@app.route('/matcher', methods=['POST'])
def matcher():
    if request.method == 'POST':
        job_description = request.form['job_description']
        resume_files = request.files.getlist('resumes')

        if not resume_files or not job_description:
            return render_template('matchresume.html', message="Please upload resumes and enter a job description.")

        resumes_data = []
        for resume_file in resume_files:
            filename = os.path.join(app.config['UPLOAD_FOLDER'], resume_file.filename)
            resume_file.save(filename)
            text = extract_text(filename)
            skills = extract_skills(text)
            resumes_data.append({
                'filename': resume_file.filename,
                'text': text,
                'skills': skills
            })

        # Vectorize job description and resumes
        texts = [job_description] + [r['text'] for r in resumes_data]
        vectorizer = TfidfVectorizer().fit_transform(texts)
        vectors = vectorizer.toarray()

        # Calculate cosine similarities
        job_vector = vectors[0]
        resume_vectors = vectors[1:]
        similarities = cosine_similarity([job_vector], resume_vectors)[0]

        # Create ranked results
        ranked_resumes = []
        for idx, score in enumerate(similarities):
            ranked_resumes.append({
                'rank': idx + 1,
                'filename': resumes_data[idx]['filename'],
                'score': round(score * 100, 2),
                'skills': resumes_data[idx]['skills']
            })

        # Sort by score
        ranked_resumes.sort(key=lambda x: x['score'], reverse=True)

        # Update ranks after sorting
        for idx, resume in enumerate(ranked_resumes):
            resume['rank'] = idx + 1

        return render_template('matchresume.html',
                               message="Analysis complete",
                               ranked_resumes=ranked_resumes,
                               top_10=ranked_resumes[:10])



@app.route('/download_excel/<filename>')
def download_excel(filename):
    return send_file(os.path.join(app.config['EXCEL_FOLDER'], filename),
                     as_attachment=True,
                     download_name='top_resumes.xlsx')


if __name__ == '__main__':
    if not os.path.exists(app.config['UPLOAD_FOLDER']):
        os.makedirs(app.config['UPLOAD_FOLDER'])
    app.run(debug=True)